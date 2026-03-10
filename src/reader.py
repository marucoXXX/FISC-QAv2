"""Agent 3: Reader - KBファイルを読み回答を生成"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import litellm

from .models import Answer, Confidence, Question

logger = logging.getLogger(__name__)

SYSTEM_PROMPT = """\
あなたはFISC（金融情報システムセンター）のセキュリティアンケートに回答する専門家です。
提供されたナレッジベース（KB）ドキュメントの内容のみに基づいて、各質問に対して正確に回答してください。

回答ルール:
1. KBドキュメントに明確な根拠がある場合のみ回答する
2. 根拠が見つからない場合は confidence="low", status="回答不可候補" とする
3. 情報が曖昧・矛盾する場合は confidence="medium", status="要確認" とする
4. 回答は具体的かつ簡潔に記述する
5. 必ず根拠となるドキュメントのセクションを明記する

出力は以下のJSON配列形式で返してください:
[
  {
    "question_no": <int>,
    "answer": "<回答テキスト>",
    "status": "対応済" | "要確認" | "回答不可候補",
    "source_references": ["<ファイル名> / <セクション>"],
    "confidence": "high" | "medium" | "low",
    "key_excerpt": "<根拠となる文章の抜粋>"
  }
]
"""


def _read_file_content(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        try:
            import fitz
            doc = fitz.open(str(path))
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            return text
        except ImportError:
            return f"[PDF file: {path.name} - PyMuPDF not installed]"
    elif suffix == ".docx":
        from docx import Document
        doc = Document(str(path))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    elif suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                line = " | ".join(str(c) for c in row if c is not None)
                if line.strip():
                    lines.append(line)
        wb.close()
        return "\n".join(lines)
    else:
        return path.read_text(encoding="utf-8", errors="replace")


def _build_user_prompt(
    questions: list[Question],
    file_contents: dict[str, str],
) -> str:
    parts = ["## ナレッジベースドキュメント\n"]
    for fname, content in file_contents.items():
        parts.append(f"### {fname}\n```\n{content}\n```\n")

    parts.append("\n## 回答対象の質問\n")
    for q in questions:
        parts.append(f"- Q{q.no} [{q.major} > {q.minor}]: {q.question}")

    parts.append("\n上記の質問すべてにJSON配列形式で回答してください。")
    return "\n".join(parts)


def run_reader(
    reader_id: str,
    questions: list[Question],
    files: list[str],
    kb_base_dir: Path,
    api_key: str,
    model: str = "claude-sonnet-4-20250514",
) -> list[Answer]:
    file_contents: dict[str, str] = {}
    for fpath in files:
        full_path = kb_base_dir / fpath
        if full_path.exists():
            file_contents[fpath] = _read_file_content(full_path)
        else:
            file_contents[fpath] = f"[ファイルが見つかりません: {fpath}]"

    user_prompt = _build_user_prompt(questions, file_contents)

    kwargs: dict = dict(
        model=model,
        max_tokens=16384,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_prompt},
        ],
        api_key=api_key or None,
    )
    if _is_openai_model(model):
        kwargs["response_format"] = {"type": "json_object"}

    response = litellm.completion(**kwargs)

    choice = response.choices[0]
    finish_reason = getattr(choice, "finish_reason", None)
    response_text = choice.message.content

    if not response_text:
        logger.warning("[Reader %s] Empty response from %s", reader_id, model)
        return [
            Answer(
                question_no=q.no,
                answer="",
                status="回答不可候補",
                confidence=Confidence.LOW.value,
                flag="LLM応答が空",
            )
            for q in questions
        ]

    if finish_reason == "length":
        logger.warning("[Reader %s] Response truncated (finish_reason=length)", reader_id)
        answers = _try_parse_truncated(response_text, questions)
        if answers is not None:
            return answers
        return [
            Answer(
                question_no=q.no,
                answer="",
                status="回答不可候補",
                confidence=Confidence.LOW.value,
                flag="LLM応答がトークン上限で切断",
            )
            for q in questions
        ]

    return _parse_reader_response(response_text, questions)


def _is_openai_model(model: str) -> bool:
    """OpenAI系モデルかどうかを判定する。"""
    model_lower = model.lower()
    return any(prefix in model_lower for prefix in ("gpt-", "o1-", "o3-", "openai/"))


def _try_parse_truncated(text: str, questions: list[Question]):
    """切断されたJSON配列から可能な限り回答を復旧する。成功時はlist[Answer]、失敗時はNone。"""
    raw = text.strip()
    # [ ... ] の範囲をまず探す（_extract_json_arrayは途中切断に弱いため直接探す）
    start = raw.find("[")
    if start >= 0:
        json_text = raw[start:]
    else:
        json_text = raw
    # 閉じ括弧が欠けている場合、末尾を補完して試行
    suffixes = ["", "]", "}]", '"}]', '"}]]']
    for suffix in suffixes:
        try:
            data = json.loads(json_text + suffix)
            if isinstance(data, dict):
                for val in data.values():
                    if isinstance(val, list):
                        data = val
                        break
                else:
                    continue
            if isinstance(data, list) and len(data) > 0:
                answers = _parse_items(data, questions, missing_flag="LLM応答がトークン上限で切断")
                logger.info("[Reader] Recovered %d/%d answers from truncated response", len(data), len(questions))
                return answers
        except (json.JSONDecodeError, KeyError, TypeError):
            continue
    return None


def _parse_items(
    data: list,
    questions: list[Question],
    missing_flag: str = "Readerから回答なし",
) -> list[Answer]:
    """パース済みのJSONリストからAnswerリストを生成する。"""
    answers: list[Answer] = []
    for item in data:
        answers.append(Answer(
            question_no=item["question_no"],
            answer=item.get("answer", ""),
            status=item.get("status", "要確認"),
            source_references=item.get("source_references", []),
            confidence=item.get("confidence", Confidence.MEDIUM.value),
            key_excerpt=item.get("key_excerpt", ""),
        ))
    # Ensure all questions have answers
    answered_nos = {a.question_no for a in answers}
    for q in questions:
        if q.no not in answered_nos:
            answers.append(Answer(
                question_no=q.no,
                answer="",
                status="回答不可候補",
                confidence=Confidence.LOW.value,
                flag=missing_flag,
            ))
    return answers


def _find_matching_bracket(text: str, open_char: str, close_char: str) -> tuple:
    """括弧深度カウンタで最初の open_char に対応する close_char の位置を見つける。

    Returns (start, end) where text[start:end] is the matched bracket content,
    or (-1, -1) if not found.
    """
    start = text.find(open_char)
    if start < 0:
        return (-1, -1)
    depth = 0
    in_string = False
    escape_next = False
    for i in range(start, len(text)):
        ch = text[i]
        if escape_next:
            escape_next = False
            continue
        if ch == '\\' and in_string:
            escape_next = True
            continue
        if ch == '"' and not escape_next:
            in_string = not in_string
            continue
        if in_string:
            continue
        if ch == open_char:
            depth += 1
        elif ch == close_char:
            depth -= 1
            if depth == 0:
                return (start, i + 1)
    return (-1, -1)


def _extract_json_array(text: str) -> str:
    """LLM応答テキストからJSON配列部分を抽出する。"""
    raw = text.strip()
    # 1) コードブロック内を探す
    if "```" in raw:
        parts = raw.split("```")
        for part in parts[1::2]:  # odd-indexed parts are inside code blocks
            stripped = part.strip()
            if stripped.startswith("json"):
                stripped = stripped[4:].strip()
            # [ で始まるJSON配列、または { で始まるオブジェクトのどちらも受け付ける
            if stripped.startswith("[") or stripped.startswith("{"):
                return stripped
    # 2) フォールバック: 生テキストから [ ... ] を括弧深度マッチングで探す
    start, end = _find_matching_bracket(raw, "[", "]")
    if start >= 0:
        return raw[start:end]
    # 3) さらにフォールバック: { ... } を括弧深度マッチングで探す（オブジェクト包みの場合）
    start, end = _find_matching_bracket(raw, "{", "}")
    if start >= 0:
        return raw[start:end]
    return raw


def _parse_reader_response(text: str, questions: list[Question]) -> list[Answer]:
    if not text or not text.strip():
        logger.warning("[Reader] Empty response text")
        return [
            Answer(
                question_no=q.no,
                answer="",
                status="回答不可候補",
                confidence=Confidence.LOW.value,
                flag="LLM応答のパースに失敗",
            )
            for q in questions
        ]

    # Step 1: 直接 json.loads を試行（response_format 対応）
    data = None
    try:
        data = json.loads(text.strip())
    except (json.JSONDecodeError, ValueError):
        pass

    # Step 2: 失敗時のみ _extract_json_array でフォールバック
    if data is None:
        json_text = _extract_json_array(text)
        try:
            data = json.loads(json_text)
        except json.JSONDecodeError:
            logger.warning("[Reader] JSON parse failed. Raw response (first 500 chars):\n%s", text[:500])
            return [
                Answer(
                    question_no=q.no,
                    answer="",
                    status="回答不可候補",
                    confidence=Confidence.LOW.value,
                    flag="LLM応答のパースに失敗",
                )
                for q in questions
            ]

    # LLMがオブジェクトで包んだ場合、中の配列を取り出す
    if isinstance(data, dict):
        for val in data.values():
            if isinstance(val, list):
                data = val
                break
        else:
            data = []

    return _parse_items(data, questions)


def answers_to_dicts(answers: list[Answer]) -> list[dict]:
    return [
        {
            "question_no": a.question_no,
            "answer": a.answer,
            "status": a.status,
            "source_references": a.source_references,
            "confidence": a.confidence,
            "key_excerpt": a.key_excerpt,
        }
        for a in answers
    ]
