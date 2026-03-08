"""Agent 1: Indexer - KBファイルの軽量インデックス生成と更新検知"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from .models import IndexEntry

# 1トークン ≈ 4文字（日本語は約2文字）として推定
_CHARS_PER_TOKEN_EN = 4
_CHARS_PER_TOKEN_JA = 2

# 要約の最大文字数
_SUMMARY_MAX_CHARS = 500

# DOCX で目次とみなすパターン
_TOC_PATTERNS = re.compile(r"^(目次|table of contents|contents)\s*$", re.IGNORECASE)


def _estimate_tokens(text: str) -> int:
    ja_count = sum(1 for c in text if ord(c) > 0x3000)
    en_count = len(text) - ja_count
    return int(ja_count / _CHARS_PER_TOKEN_JA + en_count / _CHARS_PER_TOKEN_EN)


def _extract_summary_pdf(path: Path, max_pages: int = 2) -> tuple[str, int]:
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(str(path))
        # 先頭ページからテキスト抽出（要約用）
        head_text = ""
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            head_text += page.get_text()
        # 全ページからトークン推定
        full_text = "".join(page.get_text() for page in doc)
        tokens = _estimate_tokens(full_text)
        doc.close()
        summary = head_text[:_SUMMARY_MAX_CHARS].strip()
        return summary, tokens
    except ImportError:
        # PyMuPDF not available - estimate from file size
        size = path.stat().st_size
        tokens = size // 3  # rough estimate
        return f"[PDF] {path.stem}", tokens


def _extract_summary_docx(path: Path) -> tuple[str, int]:
    from docx import Document
    doc = Document(str(path))

    headings: list[str] = []
    body_parts: list[str] = []
    in_toc = False

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        # 目次セクションの検出・スキップ
        if _TOC_PATTERNS.match(text):
            in_toc = True
            continue
        # 次の見出しが来たら目次セクション終了
        is_heading = para.style.name.startswith("Heading")
        if in_toc and is_heading:
            in_toc = False
        if in_toc:
            continue

        if is_heading:
            headings.append(text)
        else:
            body_parts.append(text)

    # 見出しを優先して要約を構成
    summary_parts: list[str] = []
    remaining = _SUMMARY_MAX_CHARS
    for h in headings:
        if remaining <= 0:
            break
        summary_parts.append(h)
        remaining -= len(h) + 1  # +1 for newline
    if remaining > 0 and body_parts:
        body_text = "\n".join(body_parts)
        summary_parts.append(body_text[:remaining])

    full_text = "\n".join(headings + body_parts)
    tokens = _estimate_tokens(full_text)
    summary = "\n".join(summary_parts).strip()
    return summary, tokens


def _extract_summary_xlsx(path: Path) -> tuple[str, int]:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    texts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join(str(c) for c in row if c is not None)
            if row_text.strip():
                texts.append(row_text)
    wb.close()
    full_text = "\n".join(texts)
    tokens = _estimate_tokens(full_text)
    summary = full_text[:_SUMMARY_MAX_CHARS].strip()
    return summary, tokens


def _extract_summary_text(path: Path, max_lines: int = 50) -> tuple[str, int]:
    text = path.read_text(encoding="utf-8", errors="replace")
    tokens = _estimate_tokens(text)
    lines = text.splitlines()[:max_lines]
    summary = "\n".join(lines)[:_SUMMARY_MAX_CHARS].strip()
    return summary, tokens


def _extract_summary(path: Path) -> tuple[str, int]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_summary_pdf(path)
    elif suffix == ".docx":
        return _extract_summary_docx(path)
    elif suffix == ".xlsx":
        return _extract_summary_xlsx(path)
    else:
        return _extract_summary_text(path)


def _llm_summarize(text: str, file_name: str, api_key: str, model: str) -> str:
    """LLM API を使って構造化サマリを生成する。"""
    import litellm
    prompt = (
        f"以下はKBドキュメント「{file_name}」の先頭部分です。\n"
        "このドキュメントの内容を100〜200トークン程度で要約してください。\n"
        "要約には以下を含めてください:\n"
        "- ドキュメントの主題・目的\n"
        "- カバーしている主要トピック（箇条書き）\n"
        "- FISC安全対策基準との関連性（あれば）\n\n"
        f"```\n{text[:3000]}\n```"
    )
    response = litellm.completion(
        model=model,
        max_tokens=512,
        messages=[{"role": "user", "content": prompt}],
        api_key=api_key or None,
    )
    return response.choices[0].message.content.strip()


def _file_modified_iso(path: Path) -> str:
    mtime = path.stat().st_mtime
    return datetime.fromtimestamp(mtime, tz=timezone.utc).isoformat()


def _category_from_path(path: Path, kb_dir: Path) -> str:
    try:
        rel = path.relative_to(kb_dir)
        return rel.parts[0] if len(rel.parts) > 1 else "other"
    except ValueError:
        return "other"


def run_indexer(
    kb_dir: Path,
    previous_index: list[dict] | None = None,
    api_key: str | None = None,
    model: str = "claude-sonnet-4-20250514",
    use_llm_summary: bool = False,
) -> list[IndexEntry]:
    kb_dir = Path(kb_dir)
    prev_map: dict[str, str] = {}
    if previous_index:
        for entry in previous_index:
            prev_map[entry["file_name"]] = entry.get("last_modified", "")

    entries: list[IndexEntry] = []
    for f in sorted(kb_dir.rglob("*")):
        if not f.is_file() or f.name.startswith("."):
            continue
        summary, tokens = _extract_summary(f)
        last_modified = _file_modified_iso(f)
        updated = True
        if f.name in prev_map and prev_map[f.name] == last_modified:
            updated = False

        # LLM要約が有効かつAPIキーがある場合、要約をLLMで生成
        if use_llm_summary and api_key and summary:
            try:
                summary = _llm_summarize(summary, f.name, api_key, model)
            except Exception:
                pass  # LLM失敗時はローカル要約をそのまま使う

        category = _category_from_path(f, kb_dir)
        entries.append(IndexEntry(
            file_name=f.name,
            path=str(f.relative_to(kb_dir.parent.parent) if "fixtures" in str(kb_dir) else f),
            category=category,
            summary=summary,
            estimated_tokens=tokens,
            last_modified=last_modified,
            updated=updated,
        ))

    return entries


def index_to_dicts(entries: list[IndexEntry]) -> list[dict]:
    return [
        {
            "file_name": e.file_name,
            "path": e.path,
            "category": e.category,
            "summary": e.summary,
            "estimated_tokens": e.estimated_tokens,
            "last_modified": e.last_modified,
            "updated": e.updated,
        }
        for e in entries
    ]


def save_index(entries: list[IndexEntry], path: Path) -> None:
    path.write_text(
        json.dumps(index_to_dicts(entries), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_index(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))
