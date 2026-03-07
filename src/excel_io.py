"""FISC-QAv2 Excel 読み書き"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import Answer, Question, ReviewNote

HEADERS = ["No.", "大分類", "小分類", "質問内容", "回答", "対応状況", "根拠ソース", "確信度", "備考"]


def read_questionnaire(path: Path) -> list[Question]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    questions: list[Question] = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    for row in rows:
        if row[0] is None:
            continue
        questions.append(Question(
            no=int(row[0]),
            major=str(row[1] or ""),
            minor=str(row[2] or ""),
            question=str(row[3] or ""),
        ))
    return questions


def write_results(
    path: Path,
    questions: list[Question],
    answers: dict[int, Answer],
    review_notes: list[ReviewNote] | None = None,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "FISC回答結果"
    ws.append(HEADERS)

    for q in questions:
        ans = answers.get(q.no)
        if ans:
            ws.append([
                q.no,
                q.major,
                q.minor,
                q.question,
                ans.answer,
                ans.status,
                ", ".join(ans.source_references) if ans.source_references else "",
                ans.confidence,
                ans.flag or "",
            ])
        else:
            ws.append([q.no, q.major, q.minor, q.question, "", "未回答", "", "", ""])

    if review_notes:
        ws2 = wb.create_sheet("レビュー指摘")
        ws2.append(["質問No.", "指摘種別", "重大度", "説明", "提案"])
        for note in review_notes:
            ws2.append([
                note.question_no,
                note.issue_type,
                note.severity,
                note.description,
                note.suggestion,
            ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def generate_output_path(output_dir: str) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(output_dir) / f"FISC回答_{ts}.xlsx"
