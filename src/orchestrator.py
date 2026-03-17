"""Agent 0: オーケストレーター - 業務フロー全体を制御"""

from __future__ import annotations

import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from .config import Config
from .excel_io import generate_output_path, read_questionnaire, write_results
from .indexer import index_to_dicts, load_index, run_indexer, save_index
from .models import Answer, Confidence, Question, ReaderAssignment
from .reader import run_reader
from .reviewer import run_reviewer
from .router import run_router


def _log(msg: str) -> None:
    print(msg, file=sys.stderr, flush=True)


def _load_past_answers(kb_dir: Path) -> dict[str, dict]:
    """過去回答をロードする。質問IDをキーとした辞書を返す。"""
    past_dir = kb_dir / "past_answers"
    if not past_dir.exists():
        return {}

    from openpyxl import load_workbook

    past: dict[str, dict] = {}
    for xlsx in sorted(past_dir.glob("*.xlsx"), reverse=True):
        if xlsx.name.startswith("~$"):
            continue
        wb = load_workbook(str(xlsx), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            qid = str(row[0])
            if qid not in past:
                past[qid] = {
                    "answer": str(row[1] or ""),
                    "evidence": str(row[2] or ""),
                    "note": str(row[3] or ""),
                    "confidence": str(row[4] or ""),
                }
        wb.close()
    return past


def _find_related_files(
    question: Question,
    index_dicts: list[dict],
    routing_map: dict[int, list[str]] | None = None,
) -> list[str]:
    """質問に関連するKBファイル名リストを返す。

    routing_map が渡された場合は Router の出力を使い、
    なければ従来の静的キーワードマッチにフォールバックする。
    """
    if routing_map and question.no in routing_map:
        return routing_map[question.no]

    from .router import _KEYWORD_MAP
    keywords = _KEYWORD_MAP.get(question.major, [])
    related = []
    for entry in index_dicts:
        stem = entry["file_name"].rsplit(".", 1)[0]
        if any(kw in stem for kw in keywords):
            related.append(entry["path"])
    return related


def _run_single_reader(
    assignment: ReaderAssignment,
    questions: list[Question],
    kb_dir: Path,
    config: Config,
) -> list[Answer]:
    """1つの Reader を最大 max_retries 回リトライ付きで実行する。"""
    assigned_qs = [q for q in questions if q.no in assignment.questions]
    last_error: Exception | None = None

    for attempt in range(config.max_reader_retries):
        try:
            return run_reader(
                reader_id=assignment.reader_id,
                questions=assigned_qs,
                files=assignment.files,
                kb_base_dir=kb_dir,
                api_key=config.api_key,
                model=config.model,
            )
        except Exception as e:
            last_error = e
            if attempt < config.max_reader_retries - 1:
                wait = 2 ** attempt
                _log(f"  Reader {assignment.reader_id} リトライ ({attempt + 1}/{config.max_reader_retries}): {e}")
                time.sleep(wait)
            else:
                _log(f"  Reader {assignment.reader_id} 失敗: {e}")

    # 全リトライ失敗 → 未回答を返す
    error_detail = str(last_error)[:200] if last_error else "unknown"
    return [
        Answer(
            question_no=q.no,
            answer="",
            status="未回答",
            confidence=Confidence.LOW.value,
            flag=f"Reader実行失敗: {error_detail}",
        )
        for q in assigned_qs
    ]


def _run_readers_parallel(
    readers: list[ReaderAssignment],
    questions: list[Question],
    kb_dir: Path,
    config: Config,
) -> list[Answer]:
    """全 Reader を ThreadPoolExecutor で並列実行し、結果を集約する。"""
    all_answers: list[Answer] = []

    with ThreadPoolExecutor(max_workers=len(readers)) as executor:
        futures = {
            executor.submit(_run_single_reader, r, questions, kb_dir, config): r
            for r in readers
        }
        for future in as_completed(futures):
            reader = futures[future]
            try:
                answers = future.result()
                all_answers.extend(answers)
                _log(f"  Reader {reader.reader_id} 完了: {len(answers)}問")
            except Exception as e:
                _log(f"  Reader {reader.reader_id} 予期しないエラー: {e}")
                assigned_qs = [q for q in questions if q.no in reader.questions]
                for q in assigned_qs:
                    all_answers.append(Answer(
                        question_no=q.no,
                        answer="",
                        status="未回答",
                        confidence=Confidence.LOW.value,
                        flag="Reader実行失敗",
                    ))

    return all_answers


def run_pipeline(
    questionnaire_path: str | Path,
    kb_dir: str | Path,
    config: Config | None = None,
) -> Path:
    config = config or Config()
    questionnaire_path = Path(questionnaire_path)
    kb_dir = Path(kb_dir)

    # [1/5] 質問票読み込み
    questions = read_questionnaire(questionnaire_path)
    _log(f"[1/5] 質問票読み込み完了: {len(questions)}問")

    # [2/5] インデックス生成 + 更新検知
    previous_index = load_index(Path(config.index_cache_path))
    index_entries = run_indexer(kb_dir, previous_index=previous_index or None)
    save_index(index_entries, Path(config.index_cache_path))
    index_dicts = index_to_dicts(index_entries)

    updated_files = [e for e in index_entries if e.updated]
    _log(
        f"[2/5] インデックス完了: KB {len(index_entries)}ファイル"
        f"（更新あり: {len(updated_files)}ファイル）"
    )

    # [3] 過去回答の事前ルーティング（LLM利用可能なら動的ルーティングで関連ファイルを特定）
    pre_routing = run_router(
        questions, index_entries, config.token_budget_per_reader,
        api_key=config.api_key, model=config.model,
    )
    # ルーティング結果から質問→ファイル名マップを構築
    routing_map: dict[int, list[str]] = {}
    for reader in pre_routing.readers:
        for qno in reader.questions:
            file_paths = [
                entry.path for entry in index_entries
                if entry.path in reader.files
            ]
            if qno not in routing_map:
                routing_map[qno] = []
            for fp in file_paths:
                if fp not in routing_map[qno]:
                    routing_map[qno].append(fp)

    past_answers = _load_past_answers(kb_dir)
    reused: dict[int, Answer] = {}
    new_questions: list[Question] = []

    for q in questions:
        related_files = _find_related_files(q, index_dicts, routing_map)
        all_unchanged = all(
            not entry.updated
            for entry in index_entries
            if entry.path in related_files
        )
        qid = f"Q{q.no:03d}"
        if all_unchanged and qid in past_answers:
            pa = past_answers[qid]
            reused[q.no] = Answer(
                question_no=q.no,
                answer=pa["answer"],
                status="過去回答採用",
                source_references=[f"past_answers ({pa.get('note', '')})"],
                confidence="past_answer",
                flag="過去回答採用",
            )
        else:
            new_questions.append(q)

    _log(f"      → 過去回答採用: {len(reused)}問 / 新規回答: {len(new_questions)}問")

    # [3/5] ルーティング（新規質問のみ再ルーティング）
    all_answers: dict[int, Answer] = dict(reused)

    if new_questions:
        routing = run_router(
            new_questions, index_entries, config.token_budget_per_reader,
            api_key=config.api_key, model=config.model,
        )
        _log(f"[3/5] ルーティング完了: Reader {len(routing.readers)}台に分配")

        # [4/5] Reader 並列実行（Fan-out → Fan-in）
        reader_results = _run_readers_parallel(
            routing.readers, new_questions, kb_dir, config,
        )

        success_count = 0
        fail_count = 0
        for ans in reader_results:
            all_answers[ans.question_no] = ans
            if ans.confidence != Confidence.LOW.value:
                success_count += 1
            else:
                fail_count += 1

        _log(
            f"[4/5] 回答生成完了: {len(new_questions)}/{len(new_questions)}問"
            f"（成功: {success_count}, 回答不可: {fail_count}）"
        )
    else:
        _log("[3/5] ルーティング: 新規回答不要（全問過去回答採用）")
        _log("[4/5] 回答生成: スキップ")

    # [5/5] レビュー
    final_answers, review_notes = run_reviewer(
        questions=questions,
        answers=all_answers,
        api_key=config.api_key,
        model=config.model,
    )

    approved = sum(
        1 for a in final_answers.values()
        if a.confidence in (Confidence.HIGH.value, "past_answer")
    )
    needs_check = len(final_answers) - approved
    _log(f"[5/5] レビュー完了: 承認推奨 {approved}問 / 要確認 {needs_check}問")

    # 出力
    output_path = generate_output_path(config.output_dir)
    write_results(output_path, questions, final_answers, review_notes or None)
    _log(f"→ 出力: {output_path}")

    return output_path
