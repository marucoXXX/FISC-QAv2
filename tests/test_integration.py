"""統合テスト: LLMモック + 実ファイルI/Oで本番さながらのパイプラインを検証する。

テストと本番の主要なギャップを埋めるテスト群:
1. Indexer → Router → Reader のパス解決が実ファイルで動作すること
2. run_pipeline 全体が LLM モックで動作し、全問に回答を返すこと
3. Reader 並列実行時に実ファイルを読み込めること
4. Web パイプラインが Excel → DB まで一貫して動作すること
5. Config 環境変数が正しく反映されること
"""

from __future__ import annotations

import json
import os
import time
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from src.config import Config
from src.indexer import run_indexer
from src.models import Answer, Confidence, Question, ReaderAssignment
from src.orchestrator import _run_readers_parallel, run_pipeline
from src.reader import _read_file_content, run_reader
from src.router import run_router, routing_to_dict


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_llm_response(text: str) -> MagicMock:
    """litellm.completion の戻り値を構築する。"""
    mock_choice = MagicMock()
    mock_choice.message.content = text
    response = MagicMock()
    response.choices = [mock_choice]
    return response


def _make_reader_response(questions: list[Question]) -> str:
    """Reader 用の正常 LLM JSON レスポンスを生成する。"""
    return json.dumps([
        {
            "question_no": q.no,
            "answer": f"Q{q.no}への回答: テスト回答テキスト",
            "status": "対応済",
            "source_references": [f"テスト根拠ドキュメント"],
            "confidence": "high",
            "key_excerpt": "テスト抜粋",
        }
        for q in questions
    ])


def _make_reviewer_response(questions: list[Question]) -> str:
    """Reviewer 用の正常 LLM JSON レスポンスを生成する。"""
    return json.dumps({
        "final_judgments": [
            {
                "question_no": q.no,
                "confidence_override": None,
                "status_override": None,
                "flag": None,
            }
            for q in questions
        ],
        "review_notes": [],
    })


# ---------------------------------------------------------------------------
# 1. パス解決テスト: Indexer → Router → Reader のファイルパスが一貫すること
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 0. KBファイル品質テスト: ファイルが意味のあるテキストを含むこと
# ---------------------------------------------------------------------------

class TestKBFileQuality:
    """KBファイルから抽出したテキストが十分な品質であることを検証する。"""

    def test_all_pdf_files_have_japanese_text(self, kb_dir: Path):
        """全PDFファイルから日本語テキストが正しく抽出できること。"""
        for pdf in sorted(kb_dir.rglob("*.pdf")):
            content = _read_file_content(pdf)
            # 最低500文字のテキストが抽出されること
            assert len(content) >= 500, (
                f"{pdf.name}: テキストが短すぎる ({len(content)}文字)"
            )
            # 日本語文字（ひらがな・カタカナ・漢字）が含まれること
            jp_chars = sum(1 for c in content if ord(c) > 0x3000)
            assert jp_chars >= 50, (
                f"{pdf.name}: 日本語文字が少なすぎる ({jp_chars}文字) — "
                "フォント問題でテキスト化が崩壊している可能性あり"
            )

    def test_all_docx_files_have_japanese_text(self, kb_dir: Path):
        """全DOCXファイルから日本語テキストが正しく抽出できること。"""
        for docx in sorted(kb_dir.rglob("*.docx")):
            content = _read_file_content(docx)
            assert len(content) >= 500, (
                f"{docx.name}: テキストが短すぎる ({len(content)}文字)"
            )
            jp_chars = sum(1 for c in content if ord(c) > 0x3000)
            assert jp_chars >= 50, (
                f"{docx.name}: 日本語文字が少なすぎる ({jp_chars}文字)"
            )

    def test_xlsx_file_has_structured_data(self, kb_dir: Path):
        """Excelファイルから構造化データが抽出できること。"""
        for xlsx in sorted(kb_dir.rglob("*.xlsx")):
            if "past_answer" in xlsx.name:
                continue
            content = _read_file_content(xlsx)
            assert len(content) >= 100, (
                f"{xlsx.name}: データが少なすぎる ({len(content)}文字)"
            )

    def test_kb_files_contain_fisc_keywords(self, kb_dir: Path):
        """KBファイルにFISC関連キーワードが含まれていること。"""
        fisc_keywords = ["FISC", "セキュリティ", "アクセス", "暗号", "認証", "監査"]
        all_content = ""
        for f in sorted(kb_dir.rglob("*")):
            if f.is_file() and not f.name.startswith(".") and "past_answer" not in f.name:
                all_content += _read_file_content(f) + "\n"

        for kw in fisc_keywords:
            assert kw in all_content, (
                f"KBファイル全体にキーワード '{kw}' が含まれていない"
            )

    def test_question_keywords_found_in_assigned_kb(self, questionnaire_path: Path, kb_dir: Path):
        """各質問の大分類キーワードが割り当てKBファイルに含まれること。"""
        from src.excel_io import read_questionnaire
        questions = read_questionnaire(questionnaire_path)
        index = run_indexer(kb_dir)
        routing = run_router(questions, index)

        # Router のファイル割り当て → 質問ごとのファイルマップ構築
        qno_files: dict[int, list[str]] = {}
        for reader in routing.readers:
            for qno in reader.questions:
                qno_files.setdefault(qno, []).extend(reader.files)

        # 大分類→期待キーワード（ファイル内容に含まれるべき語）
        major_keywords = {
            "セキュリティ管理": ["セキュリティ", "ポリシー"],
            "アクセス管理": ["アクセス", "認証"],
            "ネットワーク管理": ["ネットワーク"],
            "バックアップ管理": ["バックアップ", "復旧"],
            "インシデント対応": ["インシデント"],
            "変更管理": ["変更"],
            "教育・訓練": ["教育", "訓練"],
            "外部委託管理": ["委託"],
            "システム開発": ["開発"],
            "ログ・監視": ["ログ", "監視"],
        }

        for q in questions:
            keywords = major_keywords.get(q.major, [])
            if not keywords:
                continue
            files = qno_files.get(q.no, [])
            if not files:
                continue
            # 割り当てファイルのテキストを結合
            combined = ""
            for fpath in files:
                full_path = kb_dir / fpath
                if full_path.exists():
                    combined += _read_file_content(full_path)
            # 少なくとも1つのキーワードがファイル内容に含まれること
            found = any(kw in combined for kw in keywords)
            assert found, (
                f"Q{q.no} ({q.major}): 割り当てファイル {files} に "
                f"キーワード {keywords} のいずれも含まれていない"
            )


class TestPathResolution:
    """Indexer が生成するパスが Router/Reader で正しく使われることを検証する。"""

    def test_indexer_paths_are_kb_relative(self, kb_dir: Path):
        """Indexer が kb_dir 相対パスを生成すること。"""
        entries = run_indexer(kb_dir)
        for entry in entries:
            # パスはスラッシュを含む相対パス (e.g., "policies/security_policy.pdf")
            assert not Path(entry.path).is_absolute(), (
                f"パスが絶対パス: {entry.path}"
            )
            # kb_dir と結合して実ファイルが存在すること
            full_path = kb_dir / entry.path
            assert full_path.exists(), (
                f"kb_dir/{entry.path} が存在しない: {full_path}"
            )

    def test_router_files_match_indexer_paths(self, questionnaire_path: Path, kb_dir: Path):
        """Router の ReaderAssignment.files が Indexer の entry.path と一致すること。"""
        from src.excel_io import read_questionnaire
        questions = read_questionnaire(questionnaire_path)
        index = run_indexer(kb_dir)
        routing = run_router(questions, index)

        # Indexer のパス集合
        index_paths = {e.path for e in index}

        # Router が割り当てたファイルが全て Indexer のパスに含まれること
        for reader in routing.readers:
            for f in reader.files:
                assert f in index_paths, (
                    f"Router のファイル '{f}' が Indexer のパスにない: {index_paths}"
                )

    def test_reader_can_resolve_router_files(self, questionnaire_path: Path, kb_dir: Path):
        """Reader が Router の出力ファイルパスを kb_dir で解決できること。"""
        from src.excel_io import read_questionnaire
        questions = read_questionnaire(questionnaire_path)
        index = run_indexer(kb_dir)
        routing = run_router(questions, index)

        for reader in routing.readers:
            for fpath in reader.files:
                full_path = kb_dir / fpath
                assert full_path.exists(), (
                    f"Reader がファイルを解決できない: {full_path}"
                )
                assert full_path.is_file(), (
                    f"パスがファイルではない: {full_path}"
                )


# ---------------------------------------------------------------------------
# 2. Reader 並列実行テスト（実ファイル読み込み + LLM モック）
# ---------------------------------------------------------------------------

class TestReaderRealFiles:
    """LLM をモックしつつ、実ファイルの読み込みが正常に動作することを検証する。"""

    def test_single_reader_reads_real_files(self, questionnaire_path: Path, kb_dir: Path):
        """単一 Reader が実 KB ファイルを読み込んでLLMに渡すことを検証する。"""
        from src.excel_io import read_questionnaire
        questions = read_questionnaire(questionnaire_path)[:3]

        captured_prompt = {}

        def mock_completion(**kwargs):
            # LLM に渡されるプロンプトをキャプチャ
            messages = kwargs.get("messages", [])
            for m in messages:
                if m["role"] == "user":
                    captured_prompt["text"] = m["content"]
            return _make_llm_response(_make_reader_response(questions))

        with patch("src.reader.litellm.completion", side_effect=mock_completion):
            answers = run_reader(
                reader_id="test",
                questions=questions,
                files=["policies/security_policy.pdf"],
                kb_base_dir=kb_dir,
                api_key="fake-key",
            )

        assert len(answers) == 3
        # LLM に渡されたプロンプトにファイル内容が含まれること（"ファイルが見つかりません" ではない）
        prompt_text = captured_prompt["text"]
        assert "[ファイルが見つかりません" not in prompt_text, (
            "Reader が実ファイルを読み込めていない"
        )
        assert "security_policy.pdf" in prompt_text

    def test_reader_with_subdirectory_files(self, kb_dir: Path):
        """サブディレクトリ内のファイルが正しく読み込めること。"""
        questions = [Question(no=1, major="テスト", minor="テスト", question="テスト質問")]

        captured_prompt = {}

        def mock_completion(**kwargs):
            messages = kwargs.get("messages", [])
            for m in messages:
                if m["role"] == "user":
                    captured_prompt["text"] = m["content"]
            return _make_llm_response(_make_reader_response(questions))

        # 各サブディレクトリから1ファイルずつテスト
        test_files = [
            "policies/security_policy.pdf",
            "system_docs/incident_response.docx",
            "operations/change_management_log.xlsx",
            "regulations/outsourcing_management.pdf",
        ]

        with patch("src.reader.litellm.completion", side_effect=mock_completion):
            answers = run_reader(
                reader_id="test",
                questions=questions,
                files=test_files,
                kb_base_dir=kb_dir,
                api_key="fake-key",
            )

        assert len(answers) == 1
        prompt_text = captured_prompt["text"]
        assert "[ファイルが見つかりません" not in prompt_text
        # 全ファイルがプロンプトに含まれること
        for f in test_files:
            assert f in prompt_text, f"ファイル '{f}' がプロンプトにない"

    def test_parallel_readers_real_files(self, questionnaire_path: Path, kb_dir: Path):
        """複数 Reader が並列で実ファイルを読み込めること。"""
        from src.excel_io import read_questionnaire
        questions = read_questionnaire(questionnaire_path)
        index = run_indexer(kb_dir)
        routing = run_router(questions, index, token_budget_per_reader=5000)

        assert len(routing.readers) > 1, "テスト前提: 複数 Reader に分割されること"

        def mock_completion(**kwargs):
            messages = kwargs.get("messages", [])
            user_msg = next(m["content"] for m in messages if m["role"] == "user")
            assert "[ファイルが見つかりません" not in user_msg, (
                "並列 Reader がファイルを読み込めていない"
            )
            # Reader に割り当てられた質問を解析してレスポンスを生成
            import re
            qnos = [int(m) for m in re.findall(r"Q(\d+)\s", user_msg)]
            qs = [q for q in questions if q.no in qnos]
            return _make_llm_response(_make_reader_response(qs))

        config = Config(max_reader_retries=1, api_key="fake-key")
        with patch("src.reader.litellm.completion", side_effect=mock_completion):
            results = _run_readers_parallel(
                routing.readers, questions, kb_dir, config,
            )

        answered_nos = {a.question_no for a in results}
        expected_nos = {q.no for q in questions}
        assert answered_nos == expected_nos, (
            f"未回答: {expected_nos - answered_nos}"
        )


# ---------------------------------------------------------------------------
# 3. run_pipeline 統合テスト（LLM 全モック + 実ファイルI/O）
# ---------------------------------------------------------------------------

class TestPipelineIntegration:
    """run_pipeline を LLM モックで実行し、全問に回答が返ることを検証する。"""

    @staticmethod
    def _build_mock_completion():
        """Router/Reader/Reviewer 全ての LLM 呼び出しに対応するモック。"""
        call_count = {"n": 0}

        def mock_completion(**kwargs):
            call_count["n"] += 1
            messages = kwargs.get("messages", [])
            system_msg = ""
            user_msg = ""
            for m in messages:
                if m["role"] == "system":
                    system_msg = m["content"]
                if m["role"] == "user":
                    user_msg = m["content"]

            # Router呼び出し判定: system が空で "ルーティング" を含む
            if "ルーティングエージェント" in user_msg:
                import re
                qnos = [int(m) for m in re.findall(r"Q(\d+)\s", user_msg)]
                fnames = re.findall(r"- (\S+\.(?:pdf|docx|xlsx))", user_msg)
                routing = {
                    "routing": [
                        {"question_no": qno, "files": fnames[:2]}
                        for qno in qnos
                    ]
                }
                return _make_llm_response(json.dumps(routing))

            # Reviewer呼び出し判定
            if "品質レビュー" in system_msg:
                import re
                qnos = [int(m) for m in re.findall(r"Q(\d+)\s", user_msg)]
                qs = [Question(no=qno, major="", minor="", question="") for qno in qnos]
                return _make_llm_response(_make_reviewer_response(qs))

            # Reader呼び出し判定
            if "FISC" in system_msg and "ナレッジベース" in system_msg:
                import re
                qnos = [int(m) for m in re.findall(r"Q(\d+)\s", user_msg)]
                qs = [Question(no=qno, major="", minor="", question="") for qno in qnos]
                # ファイルが見つからないエラーがないことを確認
                assert "[ファイルが見つかりません" not in user_msg, (
                    "Reader に渡されたファイルが見つからない"
                )
                return _make_llm_response(_make_reader_response(qs))

            # デフォルト（不明な呼び出し）
            return _make_llm_response("[]")

        return mock_completion

    def test_full_pipeline_with_mocked_llm(self, questionnaire_path: Path, kb_dir: Path, tmp_path: Path):
        """LLM モックで run_pipeline を実行し、全30問に回答が返ること。"""
        config = Config(
            api_key="fake-key",
            output_dir=str(tmp_path),
            index_cache_path=str(tmp_path / ".cache.json"),
        )

        mock_fn = self._build_mock_completion()
        with patch("litellm.completion", side_effect=mock_fn):
            output_path = run_pipeline(questionnaire_path, kb_dir, config)

        assert output_path.exists()
        assert output_path.suffix == ".xlsx"

        # 出力 Excel の内容を検証
        from openpyxl import load_workbook
        wb = load_workbook(str(output_path), read_only=True)
        ws = wb["FISC回答結果"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        wb.close()

        assert len(rows) == 30, f"回答が {len(rows)} 件（期待: 30件）"

        # 各行に回答テキストが含まれること
        empty_answers = [r for r in rows if not r[4]]  # 回答列
        assert len(empty_answers) == 0, (
            f"{len(empty_answers)} 件の空回答がある"
        )

    def test_pipeline_output_has_no_file_not_found(self, questionnaire_path: Path, kb_dir: Path, tmp_path: Path):
        """パイプライン出力にファイル未発見エラーが含まれないこと。"""
        config = Config(
            api_key="fake-key",
            output_dir=str(tmp_path),
            index_cache_path=str(tmp_path / ".cache.json"),
        )

        file_not_found_errors = []

        def mock_completion(**kwargs):
            messages = kwargs.get("messages", [])
            for m in messages:
                if m["role"] == "user" and "[ファイルが見つかりません" in m["content"]:
                    file_not_found_errors.append(m["content"][:200])

            return self._build_mock_completion()(**kwargs)

        with patch("litellm.completion", side_effect=mock_completion):
            run_pipeline(questionnaire_path, kb_dir, config)

        assert len(file_not_found_errors) == 0, (
            f"ファイル未発見エラーが {len(file_not_found_errors)} 件:\n"
            + "\n".join(file_not_found_errors)
        )


# ---------------------------------------------------------------------------
# 4. Config 環境変数テスト
# ---------------------------------------------------------------------------

class TestConfigIntegration:
    """Config が環境変数から正しく読み込まれること。"""

    def test_config_reads_env_vars(self, tmp_path: Path):
        """環境変数が Config に反映されること。"""
        test_kb = str(tmp_path / "test_kb")
        os.makedirs(test_kb, exist_ok=True)

        env_overrides = {
            "FISC_KB_DIR": test_kb,
            "FISC_MODEL": "test-model",
            "FISC_TOKEN_BUDGET": "12345",
            "FISC_OUTPUT_DIR": str(tmp_path / "out"),
        }
        with patch.dict(os.environ, env_overrides):
            config = Config()

        assert config.kb_dir == test_kb
        assert config.model == "test-model"
        assert config.token_budget_per_reader == 12345
        assert config.output_dir == str(tmp_path / "out")

    def test_config_default_values(self):
        """環境変数未設定時のデフォルト値が正しいこと。"""
        env_clear = {
            "FISC_KB_DIR": "",
            "FISC_MODEL": "",
            "FISC_TOKEN_BUDGET": "",
            "FISC_OUTPUT_DIR": "",
        }
        with patch.dict(os.environ, {}, clear=False):
            # 環境変数を一時的にクリア
            for k in env_clear:
                os.environ.pop(k, None)
            config = Config()

        assert config.kb_dir == "kb"
        assert config.model == "claude-sonnet-4-20250514"
        assert config.token_budget_per_reader == 80000


# ---------------------------------------------------------------------------
# 5. Web パイプライン統合テスト（LLM モック + 実 DB）
# ---------------------------------------------------------------------------

class TestWebPipelineIntegration:
    """Web UI のパイプライン実行→DB保存まで一貫して動作することを検証する。"""

    def test_pipeline_job_writes_to_db(self, questionnaire_path: Path, kb_dir: Path, tmp_path: Path):
        """パイプラインジョブが実行されて結果がDBに保存されること。"""
        from src.web import db
        from src.web.pipeline import start_pipeline_job, get_job

        db_path = tmp_path / "test.db"
        db.init_db(db_path)

        config = Config(
            api_key="fake-key",
            output_dir=str(tmp_path / "output"),
            index_cache_path=str(tmp_path / ".cache.json"),
        )
        os.makedirs(tmp_path / "output", exist_ok=True)

        mock_fn = TestPipelineIntegration._build_mock_completion()
        with patch("litellm.completion", side_effect=mock_fn):
            job_id = start_pipeline_job(questionnaire_path, kb_dir, config, db_path)

            # ジョブ完了を待つ（最大30秒）
            for _ in range(60):
                job = get_job(job_id)
                if job and job.status in ("done", "error"):
                    break
                time.sleep(0.5)

        job = get_job(job_id)
        assert job is not None
        assert job.status == "done", f"ジョブ失敗: {job.error}"
        assert job.result_run_id is not None

        # DB に回答が保存されていることを確認
        answers = db.get_answers(db_path, job.result_run_id)
        assert len(answers) == 30, f"DB回答数: {len(answers)}（期待: 30）"
