"""E2E パイプラインテスト"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.excel_io import read_questionnaire, write_results
from src.indexer import index_to_dicts, run_indexer
from src.models import Answer, Confidence, Question, ReviewNote
from src.router import routing_to_dict, run_router


@pytest.mark.e2e
class TestE2EPipeline:
    """Indexer → Router → Reader → Reviewer の全パイプラインを検証する。"""

    def test_fixture_files_complete(self, input_dir: Path, kb_dir: Path, expected_dir: Path):
        """テストに必要な全フィクスチャが揃っていること。"""
        assert (input_dir / "questionnaire.xlsx").exists()
        assert (expected_dir / "index.json").exists()
        assert (expected_dir / "routing_map.json").exists()
        assert (expected_dir / "final_answers.json").exists()

        kb_files = list(kb_dir.rglob("*"))
        kb_file_count = sum(1 for f in kb_files if f.is_file() and not f.name.startswith("."))
        assert kb_file_count == 17

    def test_indexer_to_router_pipeline(self, questionnaire_path: Path, kb_dir: Path):
        """Indexer → Router の連携が正しく動作すること。"""
        questions = read_questionnaire(questionnaire_path)
        assert len(questions) == 30

        index = run_indexer(kb_dir)
        assert len(index) == 17

        routing = run_router(questions, index)
        result = routing_to_dict(routing)

        all_assigned = set()
        for reader in result["readers"]:
            all_assigned.update(reader["questions"])
        assert all_assigned == {q.no for q in questions}

    @pytest.mark.slow
    def test_full_pipeline(self, questionnaire_path: Path, kb_dir: Path):
        """全パイプラインを実行して回答を生成する（API必要）。"""
        pytest.importorskip("anthropic")
        import os
        if not os.environ.get("ANTHROPIC_API_KEY"):
            pytest.skip("ANTHROPIC_API_KEY not set")

        from src.orchestrator import run_pipeline
        output = run_pipeline(questionnaire_path, kb_dir)
        assert output.exists()
        assert output.suffix == ".xlsx"


class TestExcelOutput:
    """Excel 出力の検証テスト。"""

    @pytest.fixture
    def sample_data(self):
        questions = [
            Question(no=1, major="セキュリティ", minor="認証", question="MFA を導入していますか"),
            Question(no=2, major="運用", minor="監視", question="ログ監視の体制は"),
        ]
        answers = {
            1: Answer(question_no=1, answer="MFA導入済み", status="対応済",
                      confidence=Confidence.HIGH.value,
                      source_references=["security_policy.pdf / 3章"]),
            2: Answer(question_no=2, answer="", status="回答不可",
                      confidence=Confidence.LOW.value, flag="KB に該当情報なし"),
        }
        review_notes = [
            ReviewNote(question_no=2, issue_type="missing_evidence",
                       severity="high", description="根拠なし", suggestion="KB追加を推奨"),
        ]
        return questions, answers, review_notes

    def test_write_results_sheet1(self, tmp_path, sample_data):
        """Sheet1（回答結果）が正しく書き出されること。"""
        questions, answers, review_notes = sample_data
        out = tmp_path / "test_output.xlsx"
        write_results(out, questions, answers, review_notes)

        wb = load_workbook(str(out), read_only=True)
        ws = wb["FISC回答結果"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        # ヘッダー + 2データ行
        assert len(rows) == 3
        assert rows[0][0] == "No."
        assert rows[1][4] == "MFA導入済み"  # Q1 の回答
        assert rows[2][5] == "回答不可"     # Q2 のステータス

    def test_write_results_sheet2_review(self, tmp_path, sample_data):
        """Sheet2（レビュー指摘）が正しく書き出されること。"""
        questions, answers, review_notes = sample_data
        out = tmp_path / "test_output.xlsx"
        write_results(out, questions, answers, review_notes)

        wb = load_workbook(str(out), read_only=True)
        assert "レビュー指摘" in wb.sheetnames
        ws = wb["レビュー指摘"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()

        # ヘッダー + 1データ行
        assert len(rows) == 2
        assert rows[0] == ("質問No.", "指摘種別", "重大度", "説明", "提案")
        assert rows[1][0] == 2
        assert rows[1][1] == "missing_evidence"
        assert rows[1][3] == "根拠なし"
