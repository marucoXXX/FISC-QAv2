"""Tests for FISC-QAv2 Review Web UI (db + API)."""

from __future__ import annotations

import json
import tempfile
from io import BytesIO
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from src.config import Config
from src.web import db


# --- DB Layer Tests ---


@pytest.fixture
def tmp_db(tmp_path):
    db_path = tmp_path / "test.db"
    db.init_db(db_path)
    return db_path


@pytest.fixture
def sample_data():
    questions = [
        {"no": 1, "major": "セキュリティ管理", "minor": "認証", "question": "MFAの導入状況は？"},
        {"no": 2, "major": "バックアップ管理", "minor": "RPO", "question": "RPOの設定は？"},
        {"no": 3, "major": "アクセス管理", "minor": "権限", "question": "最小権限の原則は？"},
    ]
    answers = [
        {
            "question_no": 1, "answer": "MFA導入済み", "status": "対応済み",
            "source_references": ["security_policy.pdf"], "confidence": "high",
            "key_excerpt": "MFAを全ユーザに適用", "flag": None,
        },
        {
            "question_no": 2, "answer": "RPO=4時間", "status": "対応済み",
            "source_references": ["backup_policy.docx"], "confidence": "medium",
            "key_excerpt": "", "flag": None,
        },
        {
            "question_no": 3, "answer": "", "status": "回答不可",
            "source_references": [], "confidence": "low",
            "key_excerpt": "", "flag": "KB に該当情報なし",
        },
    ]
    review_notes = [
        {
            "question_no": 2, "issue_type": "weak_reference", "severity": "medium",
            "description": "根拠ソースの記述が曖昧", "suggestion": "具体的なセクション番号を追記",
        },
        {
            "question_no": 3, "issue_type": "missing_evidence", "severity": "high",
            "description": "KBに関連情報がない", "suggestion": "アクセス管理規程の追加を検討",
        },
    ]
    return questions, answers, review_notes


def test_init_db(tmp_db):
    assert tmp_db.exists()


def test_import_and_list_runs(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)
    assert run_id == 1

    runs = db.list_runs(tmp_db)
    assert len(runs) == 1
    assert runs[0]["name"] == "test_run"
    assert runs[0]["total_questions"] == 3
    assert runs[0]["pending_count"] == 3


def test_get_run(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)
    run = db.get_run(tmp_db, run_id)
    assert run is not None
    assert run["name"] == "test_run"

    assert db.get_run(tmp_db, 999) is None


def test_get_answers(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    all_answers = db.get_answers(tmp_db, run_id)
    assert len(all_answers) == 3
    assert all_answers[0]["question_no"] == 1
    assert all_answers[0]["answer"] == "MFA導入済み"
    assert all_answers[0]["source_references"] == ["security_policy.pdf"]
    assert all_answers[0]["review_status"] == "pending"


def test_get_single_answer(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    ans = db.get_answer(tmp_db, run_id, 1)
    assert ans is not None
    assert ans["confidence"] == "high"

    assert db.get_answer(tmp_db, run_id, 99) is None


def test_set_review(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    ok = db.set_review(tmp_db, run_id, 1, "approved", "LGTM")
    assert ok is True

    ans = db.get_answer(tmp_db, run_id, 1)
    assert ans["review_status"] == "approved"
    assert ans["review_comment"] == "LGTM"
    assert ans["reviewed_at"] is not None

    # Run counts updated
    run = db.get_run(tmp_db, run_id)
    assert run["approved_count"] == 1
    assert run["pending_count"] == 2


def test_bulk_set_review(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    count = db.bulk_set_review(tmp_db, run_id, "approved", [1, 2])
    assert count == 2

    run = db.get_run(tmp_db, run_id)
    assert run["approved_count"] == 2
    assert run["pending_count"] == 1


def test_bulk_set_review_all(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    count = db.bulk_set_review(tmp_db, run_id, "approved")
    assert count == 3

    run = db.get_run(tmp_db, run_id)
    assert run["approved_count"] == 3
    assert run["pending_count"] == 0


def test_get_review_notes(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    result = db.get_review_notes(tmp_db, run_id)
    assert len(result) == 2
    # Sorted by severity DESC → high first
    assert result[0]["severity"] == "high"
    assert result[0]["question_no"] == 3


def test_get_run_stats(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    db.set_review(tmp_db, run_id, 1, "approved")
    db.set_review(tmp_db, run_id, 3, "rejected")

    stats = db.get_run_stats(tmp_db, run_id)
    assert stats["total"] == 3
    assert stats["approved"] == 1
    assert stats["rejected"] == 1
    assert stats["pending"] == 1
    assert stats["conf_high"] == 1
    assert stats["conf_medium"] == 1
    assert stats["conf_low"] == 1
    assert stats["review_notes_count"] == 2


def test_filter_by_review_status(tmp_db, sample_data):
    questions, answers, notes = sample_data
    run_id = db.import_run(tmp_db, "test_run", questions, answers, notes)

    db.set_review(tmp_db, run_id, 1, "approved")

    approved = db.get_answers(tmp_db, run_id, review_status="approved")
    assert len(approved) == 1
    assert approved[0]["question_no"] == 1

    pending = db.get_answers(tmp_db, run_id, review_status="pending")
    assert len(pending) == 2


# --- API Tests ---


@pytest.fixture
def client(tmp_db):
    from starlette.testclient import TestClient
    from src.web.app import create_app

    config = Config(kb_dir="kb", api_key="test-key")
    app = create_app(tmp_db, config=config)
    return TestClient(app)


@pytest.fixture
def populated_client(client, sample_data, tmp_db):
    questions, answers, notes = sample_data
    db.import_run(tmp_db, "test_run", questions, answers, notes)
    return client


def _make_test_excel():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["No.", "大分類", "小分類", "質問内容", "回答", "対応状況", "根拠ソース", "確信度", "備考"])
    ws.append([1, "セキュリティ管理", "認証", "MFAは？", "導入済み", "対応済み", "sec.pdf", "high", ""])
    ws.append([2, "バックアップ", "RPO", "RPOは？", "4時間", "対応済み", "backup.docx", "medium", ""])

    ws2 = wb.create_sheet("レビュー指摘")
    ws2.append(["質問No.", "指摘種別", "重大度", "説明", "提案"])
    ws2.append([2, "weak_reference", "medium", "曖昧", "詳細化"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def test_api_import_excel(client):
    content = _make_test_excel()
    res = client.post("/api/runs/import", files={"file": ("test.xlsx", content)})
    assert res.status_code == 200
    data = res.json()
    assert data["run_id"] == 1
    assert "2件" in data["message"]


def test_api_import_json(client):
    payload = {
        "questions": [{"no": 1, "major": "A", "minor": "B", "question": "Q?"}],
        "answers": [{"question_no": 1, "answer": "A!", "status": "ok", "confidence": "high"}],
        "review_notes": [],
    }
    content = json.dumps(payload).encode()
    res = client.post("/api/runs/import", files={"file": ("test.json", content)})
    assert res.status_code == 200


def test_api_list_runs(populated_client):
    res = populated_client.get("/api/runs")
    assert res.status_code == 200
    runs = res.json()
    assert len(runs) == 1


def test_api_get_run(populated_client):
    res = populated_client.get("/api/runs/1")
    assert res.status_code == 200
    assert res.json()["name"] == "test_run"


def test_api_get_run_404(client):
    res = client.get("/api/runs/999")
    assert res.status_code == 404


def test_api_get_answers(populated_client):
    res = populated_client.get("/api/runs/1/answers")
    assert res.status_code == 200
    answers = res.json()
    assert len(answers) == 3


def test_api_get_answers_filtered(populated_client):
    res = populated_client.get("/api/runs/1/answers?review_status=pending")
    assert len(res.json()) == 3


def test_api_get_answer_detail(populated_client):
    res = populated_client.get("/api/runs/1/answers/1")
    assert res.status_code == 200
    data = res.json()
    assert data["answer"] == "MFA導入済み"
    assert "review_notes" in data


def test_api_set_review(populated_client):
    res = populated_client.put(
        "/api/runs/1/answers/1/review",
        json={"status": "approved", "comment": "OK"},
    )
    assert res.status_code == 200

    ans = populated_client.get("/api/runs/1/answers/1").json()
    assert ans["review_status"] == "approved"
    assert ans["review_comment"] == "OK"


def test_api_set_review_invalid_status(populated_client):
    res = populated_client.put(
        "/api/runs/1/answers/1/review",
        json={"status": "invalid"},
    )
    assert res.status_code == 400


def test_api_bulk_review(populated_client):
    res = populated_client.put(
        "/api/runs/1/bulk-review",
        json={"status": "approved", "question_nos": [1, 2]},
    )
    assert res.status_code == 200
    assert res.json()["updated"] == 2


def test_api_get_review_notes(populated_client):
    res = populated_client.get("/api/runs/1/review-notes")
    assert res.status_code == 200
    assert len(res.json()) == 2


def test_api_get_stats(populated_client):
    res = populated_client.get("/api/runs/1/stats")
    assert res.status_code == 200
    stats = res.json()
    assert stats["total"] == 3
    assert stats["pending"] == 3


def test_api_export(populated_client):
    # Approve one first
    populated_client.put("/api/runs/1/answers/1/review", json={"status": "approved"})
    res = populated_client.get("/api/runs/1/export")
    assert res.status_code == 200
    assert "spreadsheet" in res.headers["content-type"]


def test_api_delete_run(populated_client):
    res = populated_client.delete("/api/runs/1")
    assert res.status_code == 200

    res = populated_client.get("/api/runs/1")
    assert res.status_code == 404


def test_api_static_html(client):
    res = client.get("/")
    assert res.status_code == 200
    assert "FISC-QAv2" in res.text


# --- Config API Tests ---


def test_api_get_config(client):
    res = client.get("/api/config")
    assert res.status_code == 200
    data = res.json()
    assert data["kb_dir"] == "kb"
    assert data["api_key_set"] is True
    assert "model" in data
    assert "token_budget" in data


def test_api_update_config(client):
    res = client.put("/api/config", json={"kb_dir": "/new/kb", "model": "claude-haiku-4-5-20251001"})
    assert res.status_code == 200

    res = client.get("/api/config")
    data = res.json()
    assert data["kb_dir"] == "/new/kb"
    assert data["model"] == "claude-haiku-4-5-20251001"


def test_api_update_config_partial(client):
    res = client.put("/api/config", json={"token_budget": 50000})
    assert res.status_code == 200

    res = client.get("/api/config")
    assert res.json()["token_budget"] == 50000


# --- Pipeline API Tests ---


def test_api_pipeline_start(client, tmp_path):
    content = _make_test_excel()

    # Create a fake kb directory
    kb_dir = tmp_path / "kb"
    kb_dir.mkdir()

    # Update config to point to our temp kb dir
    client.put("/api/config", json={"kb_dir": str(kb_dir)})

    with patch("src.web.pipeline.start_pipeline_job", return_value="abc123") as mock_start:
        res = client.post("/api/pipeline/start", files={"file": ("test.xlsx", content)})
        assert res.status_code == 200
        data = res.json()
        assert data["job_id"] == "abc123"
        mock_start.assert_called_once()


def test_api_pipeline_start_rejects_non_xlsx(client):
    res = client.post("/api/pipeline/start", files={"file": ("test.json", b"{}")})
    assert res.status_code == 400


def test_api_pipeline_status(client):
    from src.web.pipeline import JobState, _jobs

    job = JobState(job_id="test123", status="running", progress=["step1", "step2"])
    _jobs["test123"] = job

    res = client.get("/api/pipeline/test123/status")
    assert res.status_code == 200
    data = res.json()
    assert data["status"] == "running"
    assert len(data["progress"]) == 2
    assert data["progress"][0] == "step1"

    # Cleanup
    del _jobs["test123"]


def test_api_pipeline_status_404(client):
    res = client.get("/api/pipeline/nonexistent/status")
    assert res.status_code == 404


def test_api_pipeline_progress_404(client):
    res = client.get("/api/pipeline/nonexistent/progress")
    assert res.status_code == 404


# ============================================================
# New feature tests: Bank, PastQA, CommonAnswer, Session,
# Step2-5, Export, Workflow, DB unit tests
# ============================================================


def _make_questionnaire_excel(
    question_col: str = "D",
    answer_col: str = "E",
    questions: list[str] | None = None,
):
    """Create questionnaire Excel matching bank format settings."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["No.", "大分類", "小分類", "質問内容", "回答"])
    if questions is None:
        questions = ["MFAの導入状況は？", "RPOの設定は？", "最小権限の原則は？"]
    for i, q in enumerate(questions, 1):
        ws.append([i, "分類A", "小分類B", q, ""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _make_past_answer_excel(qa_pairs: list[tuple[str, str]]):
    """Create Excel with Q&A pairs for past-answer upload (cols D=question, E=answer)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["No.", "大分類", "小分類", "質問", "回答"])
    for i, (q, a) in enumerate(qa_pairs, 1):
        ws.append([i, "", "", q, a])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --- Fixtures ---


@pytest.fixture
def bank_id(client):
    """Create a test bank and return its id."""
    res = client.post("/api/banks", json={
        "name": "テスト銀行", "code": "TEST01",
        "file_format": "xlsx", "question_col": "D", "answer_col": "E",
        "header_row": 1, "data_start_row": 2,
    })
    assert res.status_code == 200
    return res.json()["id"]


@pytest.fixture
def bank_with_past_qa(client, bank_id):
    """Bank with past Q&A uploaded."""
    content = _make_past_answer_excel([
        ("MFAの導入状況は？", "MFA導入済みです"),
        ("RPOの設定は？", "RPO=4時間に設定"),
        ("バックアップの頻度は？", "日次バックアップ"),
    ])
    res = client.post(
        f"/api/banks/{bank_id}/past-answers/upload",
        files={"file": ("past.xlsx", content)},
    )
    assert res.status_code == 200
    return bank_id


@pytest.fixture
def session_id(client, bank_id, tmp_path):
    """Create a session with extracted questions."""
    content = _make_questionnaire_excel()
    res = client.post(
        f"/api/sessions?bank_id={bank_id}",
        files={"file": ("questionnaire.xlsx", content)},
    )
    assert res.status_code == 200
    return res.json()["session_id"]


@pytest.fixture
def session_with_past_match(client, bank_with_past_qa, tmp_path):
    """Session with Step2 matching already done."""
    bank_id = bank_with_past_qa
    content = _make_questionnaire_excel()
    res = client.post(
        f"/api/sessions?bank_id={bank_id}",
        files={"file": ("questionnaire.xlsx", content)},
    )
    sid = res.json()["session_id"]

    # Run Step2 matching
    res = client.post(f"/api/sessions/{sid}/step2/match")
    assert res.status_code == 200
    return sid, bank_id


# ===== 1-1. Bank API (7 tests) =====


def test_api_create_bank(client):
    res = client.post("/api/banks", json={"name": "銀行A", "code": "BANKA"})
    assert res.status_code == 200
    assert "id" in res.json()


def test_api_create_bank_duplicate(client):
    client.post("/api/banks", json={"name": "銀行A", "code": "BANKA"})
    res = client.post("/api/banks", json={"name": "銀行A", "code": "BANKA"})
    assert res.status_code == 409


def test_api_list_banks(client, bank_id):
    res = client.get("/api/banks")
    assert res.status_code == 200
    banks = res.json()
    assert len(banks) >= 1
    assert "past_qa_count" in banks[0]


def test_api_get_bank(client, bank_id):
    res = client.get(f"/api/banks/{bank_id}")
    assert res.status_code == 200
    assert res.json()["name"] == "テスト銀行"


def test_api_get_bank_404(client):
    res = client.get("/api/banks/999")
    assert res.status_code == 404


def test_api_update_bank(client, bank_id):
    res = client.put(f"/api/banks/{bank_id}", json={"name": "更新銀行"})
    assert res.status_code == 200
    bank = client.get(f"/api/banks/{bank_id}").json()
    assert bank["name"] == "更新銀行"


def test_api_delete_bank(client, bank_id):
    res = client.delete(f"/api/banks/{bank_id}")
    assert res.status_code == 200
    res = client.get(f"/api/banks/{bank_id}")
    assert res.status_code == 404


# ===== 1-2. Past QA API (4 tests) =====


def test_api_upload_past_answers_xlsx(client, bank_id):
    content = _make_past_answer_excel([("Q1?", "A1"), ("Q2?", "A2")])
    res = client.post(
        f"/api/banks/{bank_id}/past-answers/upload",
        files={"file": ("past.xlsx", content)},
    )
    assert res.status_code == 200
    assert res.json()["count"] == 2


def test_api_upload_past_answers_invalid(client, bank_id):
    res = client.post(
        f"/api/banks/{bank_id}/past-answers/upload",
        files={"file": ("past.txt", b"plain text")},
    )
    assert res.status_code == 400


def test_api_list_past_answers(client, bank_with_past_qa):
    res = client.get(f"/api/banks/{bank_with_past_qa}/past-answers")
    assert res.status_code == 200
    assert len(res.json()) == 3


def test_api_delete_past_answer(client, bank_with_past_qa):
    items = client.get(f"/api/banks/{bank_with_past_qa}/past-answers").json()
    first_id = items[0]["id"]
    res = client.delete(f"/api/banks/{bank_with_past_qa}/past-answers/{first_id}")
    assert res.status_code == 200
    items2 = client.get(f"/api/banks/{bank_with_past_qa}/past-answers").json()
    assert len(items2) == len(items) - 1


# ===== 1-3. Common Answers API (6 tests) =====


def test_api_create_common_answer(client):
    res = client.post("/api/common-answers", json={
        "question_pattern": "MFAとは", "answer_text": "多要素認証のこと",
        "category": "セキュリティ",
    })
    assert res.status_code == 200
    cid = res.json()["id"]
    detail = client.get(f"/api/common-answers/{cid}").json()
    assert detail["question_pattern"] == "MFAとは"


def test_api_list_common_answers(client):
    client.post("/api/common-answers", json={"question_pattern": "Q1", "answer_text": "A1"})
    client.post("/api/common-answers", json={"question_pattern": "Q2", "answer_text": "A2"})
    res = client.get("/api/common-answers")
    assert res.status_code == 200
    assert len(res.json()) >= 2


def test_api_list_common_answers_search(client):
    client.post("/api/common-answers", json={
        "question_pattern": "バックアップ頻度", "answer_text": "日次",
    })
    client.post("/api/common-answers", json={
        "question_pattern": "MFA導入", "answer_text": "導入済み",
    })
    res = client.get("/api/common-answers?search=バックアップ")
    assert res.status_code == 200
    results = res.json()
    assert len(results) >= 1
    assert any("バックアップ" in r["question_pattern"] for r in results)


def test_api_get_common_answer(client):
    cid = client.post("/api/common-answers", json={
        "question_pattern": "Q", "answer_text": "A",
    }).json()["id"]
    res = client.get(f"/api/common-answers/{cid}")
    assert res.status_code == 200
    assert res.json()["answer_text"] == "A"


def test_api_update_common_answer(client):
    cid = client.post("/api/common-answers", json={
        "question_pattern": "Q", "answer_text": "A",
    }).json()["id"]
    res = client.put(f"/api/common-answers/{cid}", json={"answer_text": "Updated"})
    assert res.status_code == 200
    detail = client.get(f"/api/common-answers/{cid}").json()
    assert detail["answer_text"] == "Updated"


def test_api_delete_common_answer(client):
    cid = client.post("/api/common-answers", json={
        "question_pattern": "Q", "answer_text": "A",
    }).json()["id"]
    res = client.delete(f"/api/common-answers/{cid}")
    assert res.status_code == 200
    res = client.get(f"/api/common-answers/{cid}")
    assert res.status_code == 404


# ===== 1-4. Session API (4 tests) =====


def test_api_create_session(client, bank_id):
    content = _make_questionnaire_excel()
    res = client.post(
        f"/api/sessions?bank_id={bank_id}",
        files={"file": ("q.xlsx", content)},
    )
    assert res.status_code == 200
    data = res.json()
    assert data["question_count"] == 3
    assert "session_id" in data


def test_api_create_session_bank_not_found(client):
    content = _make_questionnaire_excel()
    res = client.post(
        "/api/sessions?bank_id=999",
        files={"file": ("q.xlsx", content)},
    )
    assert res.status_code == 404


def test_api_list_sessions(client, session_id):
    res = client.get("/api/sessions")
    assert res.status_code == 200
    sessions = res.json()
    assert len(sessions) >= 1
    assert "bank_name" in sessions[0]


def test_api_get_session(client, session_id):
    res = client.get(f"/api/sessions/{session_id}")
    assert res.status_code == 200
    data = res.json()
    assert "bank_name" in data
    assert "file_format" in data


# ===== 1-5. Step2: Past answer matching API (4 tests) =====


def test_api_step2_match(client, bank_with_past_qa, tmp_path):
    bank_id = bank_with_past_qa
    content = _make_questionnaire_excel()
    sid = client.post(
        f"/api/sessions?bank_id={bank_id}",
        files={"file": ("q.xlsx", content)},
    ).json()["session_id"]

    res = client.post(f"/api/sessions/{sid}/step2/match")
    assert res.status_code == 200
    data = res.json()
    assert "matched" in data
    assert "total" in data
    assert data["matched"] > 0


def test_api_step2_results(client, session_with_past_match):
    sid, _ = session_with_past_match
    res = client.get(f"/api/sessions/{sid}/step2/results")
    assert res.status_code == 200
    results = res.json()
    assert len(results) == 3


def test_api_step2_confirm_accept(client, session_with_past_match):
    sid, _ = session_with_past_match
    # Find a question that has a match
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    matched_q = next((q for q in questions if q["past_answer_text"]), None)
    assert matched_q is not None

    res = client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": matched_q["question_no"], "confirmed": True},
    ])
    assert res.status_code == 200
    assert res.json()["confirmed"] >= 1

    # Verify answer_source changed
    updated = client.get(f"/api/sessions/{sid}/step2/results").json()
    q = next(q for q in updated if q["question_no"] == matched_q["question_no"])
    assert q["answer_source"] == "past_match"


def test_api_step2_confirm_reject(client, session_with_past_match):
    sid, _ = session_with_past_match
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    matched_q = next((q for q in questions if q["past_answer_text"]), None)
    assert matched_q is not None

    res = client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": matched_q["question_no"], "confirmed": False},
    ])
    assert res.status_code == 200

    updated = client.get(f"/api/sessions/{sid}/step2/results").json()
    q = next(q for q in updated if q["question_no"] == matched_q["question_no"])
    assert q["answer_source"] == "pending"


# ===== 1-6. Step3: Common answer matching API (4 tests) =====


@patch("litellm.completion")
def test_api_step3_match(mock_completion, client, session_with_past_match):
    sid, _ = session_with_past_match

    # Reject all Step2 matches so questions remain pending
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": False}
        for q in questions
    ])

    # Add a common answer
    cid = client.post("/api/common-answers", json={
        "question_pattern": "MFA導入状況", "answer_text": "全社導入済み",
    }).json()["id"]

    # Mock LLM
    mock_resp = MagicMock()
    mock_resp.choices = [MagicMock()]
    mock_resp.choices[0].message.content = json.dumps([
        {"question_no": 1, "common_id": cid, "score": 0.9},
    ])
    mock_completion.return_value = mock_resp

    res = client.post(f"/api/sessions/{sid}/step3/match")
    assert res.status_code == 200
    assert res.json()["matched"] >= 1


@patch("litellm.completion")
def test_api_step3_results(mock_completion, client, session_with_past_match):
    sid, _ = session_with_past_match

    # Confirm Q1 as past_match, reject others
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    matched_q = next((q for q in questions if q["past_answer_text"]), None)
    confirm_items = []
    for q in questions:
        if q["question_no"] == matched_q["question_no"]:
            confirm_items.append({"question_no": q["question_no"], "confirmed": True})
        else:
            confirm_items.append({"question_no": q["question_no"], "confirmed": False})
    client.put(f"/api/sessions/{sid}/step2/confirm", json=confirm_items)

    # Step3 results should only show pending (unresolved)
    res = client.get(f"/api/sessions/{sid}/step3/results")
    assert res.status_code == 200
    results = res.json()
    assert all(q["answer_source"] == "pending" for q in results)


@patch("litellm.completion")
def test_api_step3_confirm(mock_completion, client, session_with_past_match):
    sid, _ = session_with_past_match

    # Reject all step2
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": False} for q in questions
    ])

    cid = client.post("/api/common-answers", json={
        "question_pattern": "MFA", "answer_text": "導入済み",
    }).json()["id"]

    mock_resp = MagicMock()
    mock_resp.choices = [MagicMock()]
    mock_resp.choices[0].message.content = json.dumps([
        {"question_no": 1, "common_id": cid, "score": 0.9},
    ])
    mock_completion.return_value = mock_resp

    client.post(f"/api/sessions/{sid}/step3/match")

    # Confirm the common match
    res = client.put(f"/api/sessions/{sid}/step3/confirm", json=[
        {"question_no": 1, "confirmed": True},
    ])
    assert res.status_code == 200
    assert res.json()["confirmed"] >= 1

    # Verify
    all_qs = client.get(f"/api/sessions/{sid}/step2/results").json()
    q1 = next(q for q in all_qs if q["question_no"] == 1)
    assert q1["answer_source"] == "common_match"


def test_api_step3_skip_all_resolved(client, session_with_past_match):
    sid, _ = session_with_past_match

    # Confirm ALL step2 matches
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": True}
        for q in questions if q["past_answer_text"]
    ])

    # Step3 match with no unresolved → matched=0
    res = client.post(f"/api/sessions/{sid}/step3/match")
    assert res.status_code == 200
    assert res.json()["matched"] == 0


# ===== 1-7. Step4: AI generation API (3 tests) =====


@patch("src.web.pipeline.start_session_pipeline_job")
def test_api_step4_generate(mock_pipeline, client, session_with_past_match):
    sid, _ = session_with_past_match
    mock_pipeline.return_value = "job123"

    # Reject all step2 so there are unresolved questions
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": False} for q in questions
    ])

    res = client.post(f"/api/sessions/{sid}/step4/generate")
    assert res.status_code == 200
    assert res.json()["job_id"] == "job123"
    mock_pipeline.assert_called_once()


def test_api_step4_generate_skip(client):
    """When all questions are resolved, Step4 should skip generation."""
    bank_id = client.post("/api/banks", json={"name": "全解決銀行", "code": "SKIP01"}).json()["id"]

    # Upload past answers matching ALL questions
    qs = ["MFAの導入状況は？", "RPOの設定は？"]
    content = _make_past_answer_excel([(q, f"回答:{q}") for q in qs])
    client.post(f"/api/banks/{bank_id}/past-answers/upload",
                files={"file": ("past.xlsx", content)})

    q_content = _make_questionnaire_excel(questions=qs)
    sid = client.post(f"/api/sessions?bank_id={bank_id}",
                      files={"file": ("q.xlsx", q_content)}).json()["session_id"]

    # Step2 match and confirm all
    client.post(f"/api/sessions/{sid}/step2/match")
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": True}
        for q in questions if q["past_answer_text"]
    ])

    res = client.post(f"/api/sessions/{sid}/step4/generate")
    assert res.status_code == 200
    data = res.json()
    assert data["skipped"] is True
    assert data["job_id"] is None


def test_api_step4_status(client, session_id):
    from src.web.pipeline import JobState, _jobs

    job = JobState(job_id="sess_job1", status="running", progress=["gen..."])
    _jobs["sess_job1"] = job

    res = client.get(f"/api/sessions/{session_id}/step4/status/sess_job1")
    assert res.status_code == 200
    assert res.json()["status"] == "running"

    del _jobs["sess_job1"]


# ===== 1-8. Step5: Final review API (6 tests) =====


def test_api_step5_summary(client, session_with_past_match):
    sid, _ = session_with_past_match

    # Confirm some matches
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": True}
        for q in questions if q["past_answer_text"]
    ])

    res = client.get(f"/api/sessions/{sid}/step5/summary")
    assert res.status_code == 200
    data = res.json()
    assert "stats" in data
    stats = data["stats"]
    assert "past_match" in stats
    assert "pending" in stats
    assert stats["total"] == 3


def test_api_edit_question(client, session_id):
    res = client.put(f"/api/sessions/{session_id}/questions/1", json={
        "answer_text": "手動で入力した回答",
    })
    assert res.status_code == 200

    questions = client.get(f"/api/sessions/{session_id}/questions").json()
    q1 = next(q for q in questions if q["question_no"] == 1)
    assert q1["answer_text"] == "手動で入力した回答"
    assert q1["answer_source"] == "manual"


def test_api_edit_question_add_to_common(client, session_id):
    res = client.put(f"/api/sessions/{session_id}/questions/1", json={
        "answer_text": "共通にも追加", "add_to_common": True,
    })
    assert res.status_code == 200

    questions = client.get(f"/api/sessions/{session_id}/questions").json()
    q1 = next(q for q in questions if q["question_no"] == 1)
    assert q1["add_to_common"] == 1


def test_api_step5_finalize(client, session_id):
    # Edit a question first
    client.put(f"/api/sessions/{session_id}/questions/1", json={
        "answer_text": "回答テスト",
    })

    res = client.put(f"/api/sessions/{session_id}/step5/finalize")
    assert res.status_code == 200
    assert res.json()["ok"] is True

    # Session status should be completed
    session = client.get(f"/api/sessions/{session_id}").json()
    assert session["status"] == "completed"


def test_api_finalize_accumulates_past_qa(client, session_id, bank_id, tmp_db):
    # Put answers on all questions
    for qno in range(1, 4):
        client.put(f"/api/sessions/{session_id}/questions/{qno}", json={
            "answer_text": f"回答{qno}",
        })

    client.put(f"/api/sessions/{session_id}/step5/finalize")

    # past_qa should now contain entries for this bank
    past = db.list_past_qa(tmp_db, bank_id)
    assert len(past) >= 3


def test_api_finalize_accumulates_common(client, session_id, tmp_db):
    # Mark Q1 for add_to_common
    client.put(f"/api/sessions/{session_id}/questions/1", json={
        "answer_text": "共通回答にも追加", "add_to_common": True,
    })
    client.put(f"/api/sessions/{session_id}/questions/2", json={
        "answer_text": "これは追加しない",
    })

    client.put(f"/api/sessions/{session_id}/step5/finalize")

    commons = db.list_common_answers(tmp_db)
    assert any("共通回答にも追加" in c["answer_text"] for c in commons)


# ===== 1-9. Export API (2 tests) =====


def test_api_export_session_xlsx(client, session_id):
    client.put(f"/api/sessions/{session_id}/questions/1", json={"answer_text": "回答"})

    res = client.get(f"/api/sessions/{session_id}/export")
    assert res.status_code == 200
    assert "spreadsheet" in res.headers["content-type"]


def test_api_export_session_fallback(client, session_id, tmp_db):
    """When source file is missing, fallback generates a new Excel."""
    # Remove the source file so it can't find the original
    with db.get_conn(tmp_db) as conn:
        conn.execute(
            "UPDATE sessions SET source_file_path = ? WHERE id = ?",
            ("/nonexistent/path.xlsx", session_id),
        )

    client.put(f"/api/sessions/{session_id}/questions/1", json={"answer_text": "回答"})
    res = client.get(f"/api/sessions/{session_id}/export")
    assert res.status_code == 200
    assert "spreadsheet" in res.headers["content-type"]


# ===== 1-10. Workflow integration tests (5 tests) =====


def test_workflow_step1_to_step2(client):
    """Bank creation → past QA upload → session creation → Step2 match."""
    # Step1: Create bank
    bank_id = client.post("/api/banks", json={
        "name": "ワークフロー銀行", "code": "WF01",
    }).json()["id"]

    # Upload past answers
    content = _make_past_answer_excel([
        ("MFAの導入状況は？", "導入済み"),
        ("RPOの設定は？", "4時間"),
    ])
    client.post(f"/api/banks/{bank_id}/past-answers/upload",
                files={"file": ("past.xlsx", content)})

    # Create session
    q_content = _make_questionnaire_excel(questions=["MFAの導入状況は？", "RPOの設定は？"])
    sid = client.post(f"/api/sessions?bank_id={bank_id}",
                      files={"file": ("q.xlsx", q_content)}).json()["session_id"]

    # Step2: match
    res = client.post(f"/api/sessions/{sid}/step2/match")
    data = res.json()
    assert data["matched"] == 2  # both should match


@patch("litellm.completion")
def test_workflow_step2_to_step3(mock_completion, client):
    """Step2 confirm → Step3 LLM matching on unresolved only."""
    bank_id = client.post("/api/banks", json={"name": "WF2銀行", "code": "WF02"}).json()["id"]

    content = _make_past_answer_excel([("MFAの導入状況は？", "導入済み")])
    client.post(f"/api/banks/{bank_id}/past-answers/upload",
                files={"file": ("past.xlsx", content)})

    q_content = _make_questionnaire_excel(questions=["MFAの導入状況は？", "バックアップ頻度は？"])
    sid = client.post(f"/api/sessions?bank_id={bank_id}",
                      files={"file": ("q.xlsx", q_content)}).json()["session_id"]

    # Step2: match and confirm Q1
    client.post(f"/api/sessions/{sid}/step2/match")
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    q1_matched = next((q for q in questions if q["past_answer_text"]), None)
    confirm_items = []
    for q in questions:
        confirm_items.append({
            "question_no": q["question_no"],
            "confirmed": bool(q["past_answer_text"]),
        })
    client.put(f"/api/sessions/{sid}/step2/confirm", json=confirm_items)

    # Add common answer for backup
    cid = client.post("/api/common-answers", json={
        "question_pattern": "バックアップ頻度", "answer_text": "日次バックアップ",
    }).json()["id"]

    # Step3: LLM mock
    mock_resp = MagicMock()
    mock_resp.choices = [MagicMock()]
    mock_resp.choices[0].message.content = json.dumps([
        {"question_no": 2, "common_id": cid, "score": 0.85},
    ])
    mock_completion.return_value = mock_resp

    res = client.post(f"/api/sessions/{sid}/step3/match")
    data = res.json()
    # Only unresolved questions should be targeted
    assert data["total"] <= 2  # at most 2
    assert data["matched"] >= 1


def test_workflow_full_past_reuse(client):
    """All questions resolved at Step2 → Step3/4 skipped → Step5 summary shows all past_match."""
    bank_id = client.post("/api/banks", json={"name": "全件マッチ", "code": "FM01"}).json()["id"]

    qs = ["MFAの導入状況は？", "RPOの設定は？"]
    content = _make_past_answer_excel([(q, f"回答:{q}") for q in qs])
    client.post(f"/api/banks/{bank_id}/past-answers/upload",
                files={"file": ("past.xlsx", content)})

    q_content = _make_questionnaire_excel(questions=qs)
    sid = client.post(f"/api/sessions?bank_id={bank_id}",
                      files={"file": ("q.xlsx", q_content)}).json()["session_id"]

    client.post(f"/api/sessions/{sid}/step2/match")
    questions = client.get(f"/api/sessions/{sid}/step2/results").json()
    client.put(f"/api/sessions/{sid}/step2/confirm", json=[
        {"question_no": q["question_no"], "confirmed": True}
        for q in questions if q["past_answer_text"]
    ])

    # Step3 → 0 matched
    res3 = client.post(f"/api/sessions/{sid}/step3/match")
    assert res3.json()["matched"] == 0

    # Step4 → skipped
    res4 = client.post(f"/api/sessions/{sid}/step4/generate")
    assert res4.json()["skipped"] is True

    # Step5 summary
    summary = client.get(f"/api/sessions/{sid}/step5/summary").json()
    assert summary["stats"]["past_match"] == 2
    assert summary["stats"]["pending"] == 0


def test_workflow_finalize_and_reuse(client, tmp_db):
    """Finalize → new session with same bank → past_qa accumulated → Step2 matches."""
    bank_id = client.post("/api/banks", json={"name": "蓄積テスト", "code": "ACC01"}).json()["id"]

    # First session with manual answers
    q_content = _make_questionnaire_excel(questions=["テスト質問1"])
    sid1 = client.post(f"/api/sessions?bank_id={bank_id}",
                       files={"file": ("q.xlsx", q_content)}).json()["session_id"]
    client.put(f"/api/sessions/{sid1}/questions/1", json={"answer_text": "回答1"})
    client.put(f"/api/sessions/{sid1}/step5/finalize")

    # Verify past_qa accumulated
    past = db.list_past_qa(tmp_db, bank_id)
    assert len(past) >= 1

    # Second session with same question
    sid2 = client.post(f"/api/sessions?bank_id={bank_id}",
                       files={"file": ("q2.xlsx", q_content)}).json()["session_id"]
    res = client.post(f"/api/sessions/{sid2}/step2/match")
    assert res.json()["matched"] >= 1


def test_workflow_export_has_answers(client, tmp_db):
    """Finalize → export contains answers."""
    bank_id = client.post("/api/banks", json={"name": "エクスポート", "code": "EXP01"}).json()["id"]

    q_content = _make_questionnaire_excel(questions=["質問A"])
    sid = client.post(f"/api/sessions?bank_id={bank_id}",
                      files={"file": ("q.xlsx", q_content)}).json()["session_id"]
    client.put(f"/api/sessions/{sid}/questions/1", json={"answer_text": "エクスポート回答"})
    client.put(f"/api/sessions/{sid}/step5/finalize")

    res = client.get(f"/api/sessions/{sid}/export")
    assert res.status_code == 200

    # Parse the returned Excel to verify answers are written
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    # Find the answer in the worksheet
    found = False
    for row in ws.iter_rows(values_only=True):
        if any("エクスポート回答" == str(cell or "") for cell in row):
            found = True
            break
    wb.close()
    assert found, "Answer not found in exported Excel"


# ===== 1-11. DB unit tests (8 tests) =====


def test_db_bank_cascade_delete(tmp_db):
    """Deleting a bank cascades to past_qa (ON DELETE CASCADE)."""
    bank_id = db.create_bank(tmp_db, "カスケード銀行", "CAS01")
    db.add_past_qa(tmp_db, bank_id, "Q?", "A!")

    db.delete_bank(tmp_db, bank_id)

    assert db.list_past_qa(tmp_db, bank_id) == []
    assert db.get_bank(tmp_db, bank_id) is None


def test_db_session_cascade_delete_questions(tmp_db):
    bank_id = db.create_bank(tmp_db, "セッション削除", "SD01")
    session_id = db.create_session(tmp_db, bank_id, "test")
    db.bulk_add_session_questions(tmp_db, session_id, [
        {"question_no": 1, "major": "A", "minor": "B", "question_text": "Q?"},
    ])

    # Delete session directly
    with db.get_conn(tmp_db) as conn:
        conn.execute("DELETE FROM sessions WHERE id = ?", (session_id,))

    questions = db.get_session_questions(tmp_db, session_id)
    assert questions == []


def test_db_list_banks_past_qa_count(tmp_db):
    b1 = db.create_bank(tmp_db, "銀行1", "B001")
    b2 = db.create_bank(tmp_db, "銀行2", "B002")
    db.add_past_qa(tmp_db, b1, "Q1", "A1")
    db.add_past_qa(tmp_db, b1, "Q2", "A2")

    banks = db.list_banks(tmp_db)
    bank1 = next(b for b in banks if b["id"] == b1)
    bank2 = next(b for b in banks if b["id"] == b2)
    assert bank1["past_qa_count"] == 2
    assert bank2["past_qa_count"] == 0


def test_db_common_answers_category_filter(tmp_db):
    db.create_common_answer(tmp_db, "Q1", "A1", category="セキュリティ")
    db.create_common_answer(tmp_db, "Q2", "A2", category="バックアップ")
    db.create_common_answer(tmp_db, "Qセキュ", "Aセキュ", category="セキュリティ")

    results = db.list_common_answers(tmp_db, category="セキュリティ")
    assert len(results) == 2
    assert all(r["category"] == "セキュリティ" for r in results)

    # Category + search combined
    results2 = db.list_common_answers(tmp_db, category="セキュリティ", search="Q1")
    assert len(results2) == 1


def test_db_update_bank_partial(tmp_db):
    bank_id = db.create_bank(tmp_db, "部分更新", "PU01", notes="初期メモ")
    ok = db.update_bank(tmp_db, bank_id, name="更新後")
    assert ok is True

    bank = db.get_bank(tmp_db, bank_id)
    assert bank["name"] == "更新後"
    assert bank["notes"] == "初期メモ"  # unchanged


def test_db_update_session_question_refs_json(tmp_db):
    bank_id = db.create_bank(tmp_db, "JSON銀行", "JS01")
    session_id = db.create_session(tmp_db, bank_id, "test")
    db.bulk_add_session_questions(tmp_db, session_id, [
        {"question_no": 1, "question_text": "Q"},
    ])

    db.update_session_question(tmp_db, session_id, 1,
                               source_references=["ref1.pdf", "ref2.docx"])

    questions = db.get_session_questions(tmp_db, session_id)
    assert questions[0]["source_references"] == ["ref1.pdf", "ref2.docx"]


def test_db_get_unresolved_only_pending(tmp_db):
    bank_id = db.create_bank(tmp_db, "未解決", "UR01")
    session_id = db.create_session(tmp_db, bank_id, "test")
    db.bulk_add_session_questions(tmp_db, session_id, [
        {"question_no": 1, "question_text": "Q1"},
        {"question_no": 2, "question_text": "Q2"},
    ])
    db.update_session_question(tmp_db, session_id, 1,
                               answer_source="past_match", answer_text="A1")

    unresolved = db.get_unresolved_questions(tmp_db, session_id)
    assert len(unresolved) == 1
    assert unresolved[0]["question_no"] == 2


def test_db_create_bank_unique_constraint(tmp_db):
    db.create_bank(tmp_db, "ユニーク", "UQ01")
    with pytest.raises(Exception, match="UNIQUE"):
        db.create_bank(tmp_db, "ユニーク", "UQ02")  # same name


# --- Bank QA Files DB Tests ---


def test_db_create_bank_qa_file(tmp_db):
    bank_id = db.create_bank(tmp_db, "QAファイル銀行", "QF01")
    qa_file_id = db.create_bank_qa_file(tmp_db, bank_id, "テストシート")
    assert qa_file_id >= 1


def test_db_list_bank_qa_files(tmp_db):
    bank_id = db.create_bank(tmp_db, "一覧銀行", "LS01")
    db.create_bank_qa_file(tmp_db, bank_id, "シートA")
    db.create_bank_qa_file(tmp_db, bank_id, "シートB")
    files = db.list_bank_qa_files(tmp_db, bank_id)
    assert len(files) == 2
    assert files[0]["qa_file_name"] == "シートA"
    assert files[1]["qa_file_name"] == "シートB"


def test_db_delete_bank_qa_file(tmp_db):
    bank_id = db.create_bank(tmp_db, "削除銀行", "DL01")
    qa_file_id = db.create_bank_qa_file(tmp_db, bank_id, "削除シート")
    assert db.delete_bank_qa_file(tmp_db, qa_file_id) is True
    assert db.list_bank_qa_files(tmp_db, bank_id) == []
    assert db.delete_bank_qa_file(tmp_db, 9999) is False


def test_db_bank_cascade_deletes_qa_files(tmp_db):
    bank_id = db.create_bank(tmp_db, "カスケード銀行", "CS01")
    db.create_bank_qa_file(tmp_db, bank_id, "カスケードシート")
    db.delete_bank(tmp_db, bank_id)
    assert db.list_bank_qa_files(tmp_db, bank_id) == []


def test_db_list_banks_qa_file_count(tmp_db):
    bank_id = db.create_bank(tmp_db, "カウント銀行", "CT01")
    db.create_bank_qa_file(tmp_db, bank_id, "シート1")
    db.create_bank_qa_file(tmp_db, bank_id, "シート2")
    banks = db.list_banks(tmp_db)
    bank = next(b for b in banks if b["id"] == bank_id)
    assert bank["qa_file_count"] == 2


# --- Bank QA Files API Tests ---


def test_api_create_bank_qa_file(client, bank_id):
    res = client.post(f"/api/banks/{bank_id}/qa-files", json={"qa_file_name": "APIシート"})
    assert res.status_code == 200
    assert "id" in res.json()


def test_api_list_bank_qa_files(client, bank_id):
    client.post(f"/api/banks/{bank_id}/qa-files", json={"qa_file_name": "シートX"})
    client.post(f"/api/banks/{bank_id}/qa-files", json={"qa_file_name": "シートY"})
    res = client.get(f"/api/banks/{bank_id}/qa-files")
    assert res.status_code == 200
    assert len(res.json()) == 2


def test_api_delete_bank_qa_file(client, bank_id):
    res = client.post(f"/api/banks/{bank_id}/qa-files", json={"qa_file_name": "削除対象"})
    qa_file_id = res.json()["id"]
    res = client.delete(f"/api/banks/{bank_id}/qa-files/{qa_file_id}")
    assert res.status_code == 200
    assert res.json()["ok"] is True


def test_api_create_bank_qa_file_bank_not_found(client):
    res = client.post("/api/banks/9999/qa-files", json={"qa_file_name": "存在しない"})
    assert res.status_code == 404


# --- Seed Script Tests ---


def test_seed_banks(tmp_db):
    from scripts.seed_banks import seed_banks
    result = seed_banks(tmp_db)
    assert result["created_banks"] == 25
    assert result["created_files"] == 29
    banks = db.list_banks(tmp_db)
    assert len(banks) == 25


def test_seed_banks_idempotent(tmp_db):
    from scripts.seed_banks import seed_banks
    seed_banks(tmp_db)
    result = seed_banks(tmp_db)
    assert result["created_banks"] == 0
    assert result["skipped_banks"] == 25
    assert result["created_files"] == 0
    banks = db.list_banks(tmp_db)
    assert len(banks) == 25
